import csv

from engine.file_writer import append_row
from engine.data_source import DataSource


def _read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def test_append_row_csv_writes_header_then_rows(tmp_path):
    p = tmp_path / "out.csv"
    append_row(str(p), ["MAT-001", "10"], header=["CODE", "QTY"])
    append_row(str(p), ["MAT-002", "5"], header=["CODE", "QTY"])
    rows = _read_csv(str(p))
    assert rows == [["CODE", "QTY"], ["MAT-001", "10"], ["MAT-002", "5"]]


def test_append_row_csv_header_only_once(tmp_path):
    p = tmp_path / "out.csv"
    for i in range(3):
        append_row(str(p), [f"v{i}"], header=["H"])
    rows = _read_csv(str(p))
    assert rows == [["H"], ["v0"], ["v1"], ["v2"]]


def test_append_row_creates_parent_dir(tmp_path):
    p = tmp_path / "sub" / "out.csv"
    append_row(str(p), ["a"], header=None)
    assert p.exists()
    assert _read_csv(str(p)) == [["a"]]


def test_append_row_xlsx_roundtrip(tmp_path):
    from openpyxl import load_workbook
    p = tmp_path / "out.xlsx"
    append_row(str(p), ["MAT-001", "10"], header=["CODE", "QTY"])
    append_row(str(p), ["MAT-002", "5"])
    ws = load_workbook(str(p)).active
    got = [[c.value for c in row] for row in ws.iter_rows()]
    assert got == [["CODE", "QTY"], ["MAT-001", "10"], ["MAT-002", "5"]]


def test_read_headers_csv(tmp_path):
    p = tmp_path / "h.csv"
    p.write_text("MATERIAL_CODE,QTY,WORK_STATION\nMAT-001,10,WS-A01\n", encoding="utf-8")
    assert DataSource.read_headers(str(p)) == ["MATERIAL_CODE", "QTY", "WORK_STATION"]


def test_read_headers_xlsx(tmp_path):
    from openpyxl import Workbook
    p = tmp_path / "h.xlsx"
    wb = Workbook()
    wb.active.append(["CODE", "QTY"])
    wb.active.append(["MAT-001", "10"])
    wb.save(str(p))
    assert DataSource.read_headers(str(p)) == ["CODE", "QTY"]


def test_read_headers_missing_file_returns_empty():
    assert DataSource.read_headers("does/not/exist.csv") == []
    assert DataSource.read_headers("") == []


def test_datasource_reads_xlsx(tmp_path):
    from openpyxl import Workbook
    p = tmp_path / "src.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["MATERIAL_CODE", "QTY"])
    ws.append(["MAT-001", "10"])
    ws.append(["MAT-002", "5"])
    wb.save(str(p))

    ds = DataSource({}, str(p))
    ds.next_row()
    assert ds.resolve("{csv.MATERIAL_CODE}") == "MAT-001"
    ds.next_row()
    assert ds.resolve("{csv.QTY}") == "5"
