"""
เขียน/ต่อแถวข้อมูลลงไฟล์โดยตรง (ไม่ผ่านการพิมพ์หน้าจอ)
รองรับ .csv และ .xlsx — ใช้กับ action write_row
"""
import csv
import os

from engine.logger import get_logger


class FileWriteError(Exception):
    pass


def append_row(path: str, values: list, header: list = None):
    """ต่อหนึ่งแถวลงท้ายไฟล์ path
    - ถ้าไฟล์ยังไม่มี/ว่าง และมี header → เขียน header เป็นแถวแรกก่อน
    - .xlsx/.xlsm → ใช้ openpyxl, อื่นๆ → csv
    """
    ext = os.path.splitext(path)[1].lower()
    values = ["" if v is None else str(v) for v in values]
    try:
        if ext in (".xlsx", ".xlsm"):
            _append_xlsx(path, values, header)
        else:
            _append_csv(path, values, header)
        get_logger().info(f"write_row → {path}: {values}")
    except FileWriteError:
        raise
    except Exception as e:
        raise FileWriteError(f"เขียนไฟล์ '{path}' ไม่สำเร็จ: {e}") from e


def _append_csv(path: str, values: list, header: list = None):
    new_file = (not os.path.exists(path)) or os.path.getsize(path) == 0
    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    with open(path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if new_file and header:
            writer.writerow(header)
        writer.writerow(values)


def append_error_log(path: str, row_num: int, error_msg: str):
    """บันทึกแถวที่พลาดลงไฟล์ error log — ใช้กับ on_row_error=recover"""
    from datetime import datetime
    values = [row_num, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), error_msg]
    header = ["row_num", "timestamp", "error"]
    append_row(path, values, header)


def read_error_log_row_nums(path: str) -> set[int]:
    """อ่านหมายเลขแถวจาก error log — คืน set[int] ของ row_num ที่พลาด"""
    try:
        import pandas as pd
        df = pd.read_csv(path, dtype=str)
        if "row_num" not in df.columns:
            return set()
        result = set()
        for v in df["row_num"].dropna():
            s = str(v).strip()
            if s.lstrip("-").isdigit():
                result.add(int(float(s)))
        return result
    except Exception as e:
        get_logger().warning(f"read_error_log_row_nums('{path}'): {e}")
        return set()


def annotate_csv_row_error(csv_path: str, row_num: int, error_msg: str,
                           step_info: str, screenshot_path: str = ""):
    """เขียนข้อมูล error ลงใน row ที่กำหนดของ CSV/XLSX (in-place)
    เพิ่มคอลัมน์ error_status / error_step / error_message / error_screenshot ถ้ายังไม่มี
    row_num: 1-indexed (นับเฉพาะแถวข้อมูล ไม่นับ header)
    """
    import pandas as pd
    ext = os.path.splitext(csv_path)[1].lower()
    try:
        if ext in (".xlsx", ".xls", ".xlsm"):
            df = pd.read_excel(csv_path, dtype=str)
        else:
            df = pd.read_csv(csv_path, dtype=str)
    except Exception as e:
        get_logger().warning(f"annotate_csv_row_error: อ่าน '{csv_path}' ไม่สำเร็จ: {e}")
        return

    idx = row_num - 1
    if idx < 0 or idx >= len(df):
        get_logger().warning(f"annotate_csv_row_error: row_num {row_num} out of range (len={len(df)})")
        return

    for col in ("error_status", "error_step", "error_message", "error_screenshot"):
        if col not in df.columns:
            df[col] = ""

    df = df.fillna("")
    df.loc[idx, "error_status"] = "error"
    df.loc[idx, "error_step"] = step_info
    df.loc[idx, "error_message"] = error_msg
    df.loc[idx, "error_screenshot"] = screenshot_path

    try:
        if ext in (".xlsx", ".xlsm"):
            df.to_excel(csv_path, index=False)
        else:
            df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        get_logger().info(f"annotate_csv_row_error: row {row_num} → '{csv_path}'")
    except Exception as e:
        get_logger().warning(f"annotate_csv_row_error: เขียน '{csv_path}' ไม่สำเร็จ: {e}")


def _append_xlsx(path: str, values: list, header: list = None):
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as e:
        raise FileWriteError(f"ต้องติดตั้ง openpyxl เพื่อเขียน .xlsx: {e}") from e

    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        if header:
            ws.append(list(header))

    ws.append(values)
    wb.save(path)
